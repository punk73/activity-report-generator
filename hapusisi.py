import openpyxl

# Memuat workbook dan sheet
workbook = openpyxl.load_workbook('LAPORAN PUBLIKASI SOSMED.xlsx')
sheet = workbook.active

# Menghapus isi sel dari A61:J112
for row in sheet.iter_rows(min_row=61, max_row=112, min_col=1, max_col=10):
    for cell in row:
        if isinstance(cell, openpyxl.cell.MergedCell):
            # Mengakses sel utama dari merged cell
            main_cell = sheet.cell(row=cell.row, column=cell.column)
            if main_cell.value is not None:
                main_cell.value = None  # Menghapus isi sel utama
        else:
            cell.value = None  # Menghapus isi sel jika bukan merged cell

# Simpan workbook
workbook.save('LAPORAN.xlsx')
